import openpyxl
import os
print(os.getcwd())
wb = openpyxl.load_workbook('модели.xlsx')
sheets = wb.get_sheet_names
model_file = open('models.py','w', encoding='utf-8')
model_file.write('from django.db import models\n')
for sheet_name in sheets():
    sheet = wb.get_sheet_by_name(sheet_name)  
    model_file.write('class ' + sheet_name + ' (models.Model):')
    for row_of_cell_objects in sheet['A1':'N'+ str(sheet.max_row)]:
        s=list()
        for cell_obj in row_of_cell_objects:
            if cell_obj.row != 1:
                if sheet.cell(row=1, column=cell_obj.column).value == 'Наименование':
                    s.append('\t'+cell_obj.value + ' = models.')
                if sheet.cell(row=1, column=cell_obj.column).value == 'тип':
                    s.append(cell_obj.value +'(')
                if sheet.cell(row=1, column=cell_obj.column).value == 'main_table' and cell_obj.value != None:
                    s.append(cell_obj.value+', ')
                if sheet.cell(row=1, column=cell_obj.column).value == 'Имя':
                    s.append('verbose_name = "'+ cell_obj.value+'", ')
                if sheet.cell(row=1, column=cell_obj.column).value == 'unique' and cell_obj.value != None:
                    s.append('unique = True, ')
                if sheet.cell(row=1, column=cell_obj.column).value == 'primary_key' and cell_obj.value != None:
                    s.append('primary_key = True, ')
                if sheet.cell(row=1, column=cell_obj.column).value == 'db_index' and cell_obj.value != None:
                    s.append('db_index = True, ')
                if sheet.cell(row=1, column=cell_obj.column).value == 'max_length' and cell_obj.value != None:
                    s.append('max_length = 100, ')
                if sheet.cell(row=1, column=cell_obj.column).value == 'null' and cell_obj.value != None:
                    s.append('null = True, ')
                if sheet.cell(row=1, column=cell_obj.column).value == 'blank' and cell_obj.value != None:
                    s.append('blank = True, ')
                if sheet.cell(row=1, column=cell_obj.column).value == 'on_delete' and cell_obj.value != None:
                    s.append('on_delete = ' + cell_obj.value + ', ')
                if sheet.cell(row=1, column=cell_obj.column).value == 'choices' and cell_obj.value != None:
                    s.append('choices = ' + cell_obj.value + ', ')
                if sheet.cell(row=1, column=cell_obj.column).value == 'default' and cell_obj.value != None:
                    s.append('default = ' + str(cell_obj.value) + ', ')
                if sheet.cell(row=1, column=cell_obj.column).value == 'related_name' and cell_obj.value != None:
                    s.append('related_name = ' + "'" + cell_obj.value + "'" + ', ')
        s1= ''.join(s)[:-2]+')\n'
        s1= s1[1:]
        model_file.write('\t'+s1)
    model_file.write('\n')